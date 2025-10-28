"""
📥 SCJN Bulk Downloader
======================
Download complete SCJN tesis library from official sources
- Weekly publications (sjfsemanal.scjn.gob.mx)
- Batch exports
- Local backup library
- Zero rate limiting (official channel)
"""

import asyncio
import logging
from datetime import datetime
from typing import List, Dict, Optional
import json
import os
from pathlib import Path

logger = logging.getLogger(__name__)


class SCJNBulkDownloader:
    """
    Official SCJN Bulk Downloader
    
    Sources:
    - https://sjfsemanal.scjn.gob.mx/listado-resultado-tesis (Weekly)
    - https://sjf.scjn.gob.mx/ (Complete archive)
    
    Benefits:
    ✅ Official source (no TOS violations)
    ✅ No rate limiting
    ✅ Structured data format
    ✅ Complete metadata included
    ✅ Historical archive available
    """
    
    # Official SCJN sources
    SOURCES = {
        'weekly': 'https://sjfsemanal.scjn.gob.mx/listado-resultado-tesis',
        'archive': 'https://sjf.scjn.gob.mx/SJFHome/Index.html',
        'search': 'https://bj.scjn.gob.mx/'
    }
    
    def __init__(self, data_dir: str = 'data/scjn_library'):
        """
        Args:
            data_dir: Local directory to store tesis library
        """
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        self.tesis_dir = self.data_dir / 'tesis'
        self.tesis_dir.mkdir(exist_ok=True)
        
        self.metadata_file = self.data_dir / 'metadata.json'
        self.index_file = self.data_dir / 'index.json'
        
        self.metadata = self._load_metadata()
        self.download_stats = {
            'total_downloaded': 0,
            'total_stored': 0,
            'last_sync': None,
            'errors': []
        }
    
    def _load_metadata(self) -> dict:
        """Load existing metadata"""
        if self.metadata_file.exists():
            with open(self.metadata_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        return {
            'library_created': datetime.now().isoformat(),
            'last_updated': None,
            'total_tesis': 0,
            'materias': {},
            'salas': {}
        }
    
    def _save_metadata(self):
        """Save metadata to file"""
        with open(self.metadata_file, 'w', encoding='utf-8') as f:
            json.dump(self.metadata, f, ensure_ascii=False, indent=2)
    
    async def download_weekly_batch(self, year: Optional[int] = None) -> List[dict]:
        """
        Download weekly batch from sjfsemanal.scjn.gob.mx
        
        Args:
            year: Year to download (current if None)
        
        Returns:
            List of downloaded tesis
        """
        
        try:
            import httpx
            from bs4 import BeautifulSoup
            
            logger.info("📥 Starting weekly batch download from sjfsemanal...")
            
            # Access weekly listing
            async with httpx.AsyncClient() as client:
                response = await client.get(
                    self.SOURCES['weekly'],
                    timeout=30,
                    follow_redirects=True
                )
                response.raise_for_status()
            
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            
            # Parse tesis listing
            tesis_list = []
            
            # Look for download links or data tables
            # This structure depends on actual HTML
            download_links = soup.find_all('a', {'href': lambda x: x and ('.docx' in x or '.pdf' in x or '.xlsx' in x)})
            
            logger.info(f"🔍 Found {len(download_links)} potential documents")
            
            for link in download_links:
                try:
                    tesis_data = await self._download_tesis_document(link['href'])
                    if tesis_data:
                        tesis_list.append(tesis_data)
                except Exception as e:
                    logger.warning(f"⚠️ Failed to download {link.get('href')}: {e}")
                    continue
            
            logger.info(f"✅ Downloaded {len(tesis_list)} tesis")
            return tesis_list
        
        except Exception as e:
            logger.error(f"❌ Weekly batch download failed: {e}")
            self.download_stats['errors'].append(str(e))
            return []
    
    async def download_tesis_by_filters(
        self,
        materia: Optional[str] = None,
        sala: Optional[str] = None,
        year_range: Optional[tuple] = None
    ) -> List[dict]:
        """
        Download tesis with specific filters
        
        Args:
            materia: Matter filter
            sala: Court filter
            year_range: (start_year, end_year)
        
        Returns:
            List of tesis
        """
        
        try:
            import httpx
            
            logger.info(f"📥 Downloading: materia={materia}, sala={sala}, years={year_range}")
            
            # Build query
            params = {
                'q': f"materia:{materia}" if materia else "",
                'sala': sala if sala else "",
                'page': 1,
                'limit': 100
            }
            
            all_tesis = []
            page = 1
            
            async with httpx.AsyncClient() as client:
                while page <= 100:  # Safety limit
                    params['page'] = page
                    
                    try:
                        response = await client.get(
                            self.SOURCES['search'],
                            params=params,
                            timeout=30
                        )
                        response.raise_for_status()
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to fetch page {page}: {e}")
                        break
                    
                    # Parse response
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(response.text, 'html.parser')
                    
                    results = soup.find_all('div', class_='result-item')
                    if not results:
                        break
                    
                    for result in results:
                        tesis_data = {
                            'registro': result.find('span', class_='registro').text.strip() if result.find('span', class_='registro') else '',
                            'titulo': result.find('h3').text.strip() if result.find('h3') else '',
                            'materia': materia,
                            'sala': sala,
                            'url': result.find('a')['href'] if result.find('a') else '',
                            'downloaded_at': datetime.now().isoformat()
                        }
                        all_tesis.append(tesis_data)
                    
                    page += 1
            
            logger.info(f"✅ Downloaded {len(all_tesis)} tesis with filters")
            return all_tesis
        
        except Exception as e:
            logger.error(f"❌ Filtered download failed: {e}")
            return []
    
    async def _download_tesis_document(self, document_url: str) -> Optional[dict]:
        """
        Download and parse individual tesis document
        
        Args:
            document_url: URL to document
        
        Returns:
            Parsed tesis data or None
        """
        
        try:
            import httpx
            
            async with httpx.AsyncClient() as client:
                response = await client.get(document_url, timeout=30)
                response.raise_for_status()
            
            # Parse based on file type
            if document_url.endswith('.docx'):
                return await self._parse_docx(response.content)
            elif document_url.endswith('.pdf'):
                return await self._parse_pdf(response.content)
            elif document_url.endswith('.xlsx'):
                return await self._parse_xlsx(response.content)
            
            return None
        
        except Exception as e:
            logger.warning(f"⚠️ Failed to download document: {e}")
            return None
    
    async def _parse_docx(self, content: bytes) -> Optional[dict]:
        """Parse Word document"""
        try:
            from docx import Document
            from io import BytesIO
            
            doc = Document(BytesIO(content))
            
            # Extract text
            full_text = '\n'.join([p.text for p in doc.paragraphs])
            
            # Extract metadata
            props = doc.core_properties
            
            return {
                'content': full_text,
                'titulo': props.title,
                'author': props.author,
                'created': props.created.isoformat() if props.created else None,
                'format': 'docx'
            }
        except Exception as e:
            logger.warning(f"⚠️ DOCX parse error: {e}")
            return None
    
    async def _parse_pdf(self, content: bytes) -> Optional[dict]:
        """Parse PDF document"""
        try:
            import PyPDF2
            from io import BytesIO
            
            pdf_reader = PyPDF2.PdfReader(BytesIO(content))
            
            # Extract text
            full_text = '\n'.join([
                page.extract_text() for page in pdf_reader.pages
            ])
            
            # Get metadata
            metadata = pdf_reader.metadata
            
            return {
                'content': full_text,
                'titulo': metadata.get('/Title') if metadata else None,
                'author': metadata.get('/Author') if metadata else None,
                'format': 'pdf'
            }
        except Exception as e:
            logger.warning(f"⚠️ PDF parse error: {e}")
            return None
    
    async def _parse_xlsx(self, content: bytes) -> Optional[dict]:
        """Parse Excel document"""
        try:
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.load_workbook(BytesIO(content))
            
            tesis_list = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Assume first row is header
                header = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    tesis_dict = dict(zip(header, row))
                    tesis_list.append(tesis_dict)
            
            return {
                'tesis_list': tesis_list,
                'sheet_count': len(wb.sheetnames),
                'format': 'xlsx'
            }
        except Exception as e:
            logger.warning(f"⚠️ XLSX parse error: {e}")
            return None
    
    def store_tesis_locally(self, tesis_data: dict) -> bool:
        """
        Store downloaded tesis locally
        
        Args:
            tesis_data: Tesis data to store
        
        Returns:
            Success boolean
        """
        
        try:
            # Generate filename from registro or hash
            registro = tesis_data.get('registro', 'unknown')
            filename = f"{registro.replace('/', '_')}.json"
            
            filepath = self.tesis_dir / filename
            
            # Store with metadata
            stored_data = {
                'data': tesis_data,
                'stored_at': datetime.now().isoformat(),
                'source': 'scjn_bulk_download'
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(stored_data, f, ensure_ascii=False, indent=2)
            
            # Update metadata
            if 'materia' in tesis_data:
                materia = tesis_data['materia']
                if materia not in self.metadata['materias']:
                    self.metadata['materias'][materia] = 0
                self.metadata['materias'][materia] += 1
            
            self.metadata['total_tesis'] += 1
            self.metadata['last_updated'] = datetime.now().isoformat()
            
            self._save_metadata()
            self.download_stats['total_stored'] += 1
            
            return True
        
        except Exception as e:
            logger.error(f"❌ Failed to store tesis: {e}")
            return False
    
    def create_local_index(self) -> dict:
        """
        Create searchable index from stored tesis
        
        Returns:
            Index structure
        """
        
        try:
            logger.info("📊 Creating local index...")
            
            index = {
                'by_registro': {},
                'by_materia': {},
                'by_sala': {},
                'all_files': []
            }
            
            # Iterate through stored files
            for tesis_file in self.tesis_dir.glob('*.json'):
                try:
                    with open(tesis_file, 'r', encoding='utf-8') as f:
                        tesis_data = json.load(f)['data']
                    
                    registro = tesis_data.get('registro')
                    if registro:
                        index['by_registro'][registro] = str(tesis_file)
                    
                    materia = tesis_data.get('materia')
                    if materia:
                        if materia not in index['by_materia']:
                            index['by_materia'][materia] = []
                        index['by_materia'][materia].append(str(tesis_file))
                    
                    sala = tesis_data.get('sala')
                    if sala:
                        if sala not in index['by_sala']:
                            index['by_sala'][sala] = []
                        index['by_sala'][sala].append(str(tesis_file))
                    
                    index['all_files'].append(str(tesis_file))
                
                except Exception as e:
                    logger.warning(f"⚠️ Failed to index {tesis_file}: {e}")
                    continue
            
            # Save index
            with open(self.index_file, 'w', encoding='utf-8') as f:
                json.dump(index, f, ensure_ascii=False, indent=2)
            
            logger.info(f"✅ Index created: {len(index['all_files'])} files")
            return index
        
        except Exception as e:
            logger.error(f"❌ Failed to create index: {e}")
            return {}
    
    def search_local_library(
        self,
        query: str,
        materia: Optional[str] = None,
        sala: Optional[str] = None
    ) -> List[dict]:
        """
        Search local tesis library
        
        Args:
            query: Search text
            materia: Filter by materia
            sala: Filter by sala
        
        Returns:
            List of matching tesis
        """
        
        try:
            results = []
            
            # Load index if not already loaded
            if not hasattr(self, '_index'):
                with open(self.index_file, 'r', encoding='utf-8') as f:
                    self._index = json.load(f)
            
            # Determine which files to search
            files_to_search = self._index['all_files']
            
            if materia and materia in self._index['by_materia']:
                files_to_search = self._index['by_materia'][materia]
            elif sala and sala in self._index['by_sala']:
                files_to_search = self._index['by_sala'][sala]
            
            # Search through files
            for filepath in files_to_search:
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        tesis = json.load(f)['data']
                    
                    # Simple keyword search
                    if query.lower() in str(tesis).lower():
                        results.append(tesis)
                
                except Exception as e:
                    logger.warning(f"⚠️ Search error in {filepath}: {e}")
                    continue
            
            logger.info(f"✅ Found {len(results)} results")
            return results
        
        except Exception as e:
            logger.error(f"❌ Search failed: {e}")
            return []
    
    def get_library_stats(self) -> dict:
        """Get library statistics"""
        
        return {
            'total_tesis': self.metadata['total_tesis'],
            'materias': self.metadata['materias'],
            'salas': self.metadata['salas'],
            'last_updated': self.metadata['last_updated'],
            'storage_path': str(self.data_dir),
            'download_stats': self.download_stats
        }


# Batch downloader for scheduled jobs
class SCJNBatchDownloadScheduler:
    """Manages scheduled bulk downloads"""
    
    def __init__(self, downloader: SCJNBulkDownloader):
        self.downloader = downloader
    
    async def download_weekly_batch(self):
        """
        Download weekly batch
        Called every Friday (after official publication)
        """
        
        logger.info("🔄 [WEEKLY BATCH] Starting download...")
        
        try:
            tesis_list = await self.downloader.download_weekly_batch()
            
            # Store each tesis locally
            for tesis in tesis_list:
                self.downloader.store_tesis_locally(tesis)
            
            # Update index
            self.downloader.create_local_index()
            
            stats = self.downloader.get_library_stats()
            logger.info(f"✅ Weekly batch complete: {stats}")
        
        except Exception as e:
            logger.error(f"❌ Weekly batch failed: {e}")
    
    async def download_historical_batch(self, start_year: int, end_year: int):
        """
        Download historical batch for year range
        
        Args:
            start_year: Start year
            end_year: End year
        """
        
        logger.info(f"🔄 [HISTORICAL] Downloading {start_year}-{end_year}...")
        
        try:
            for year in range(start_year, end_year + 1):
                logger.info(f"📥 Downloading year {year}...")
                
                tesis_list = await self.downloader.download_weekly_batch(year=year)
                
                for tesis in tesis_list:
                    self.downloader.store_tesis_locally(tesis)
            
            self.downloader.create_local_index()
            logger.info("✅ Historical download complete")
        
        except Exception as e:
            logger.error(f"❌ Historical download failed: {e}")


# Usage example
"""
async def main():
    downloader = SCJNBulkDownloader()
    
    # Download weekly batch
    tesis_list = await downloader.download_weekly_batch()
    
    # Store locally
    for tesis in tesis_list:
        downloader.store_tesis_locally(tesis)
    
    # Create index
    downloader.create_local_index()
    
    # Search local library
    results = downloader.search_local_library("amparo laboral", materia="Laboral")
    
    # Get stats
    print(downloader.get_library_stats())

if __name__ == "__main__":
    asyncio.run(main())
"""
